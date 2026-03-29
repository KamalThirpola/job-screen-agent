from playwright.sync_api import sync_playwright
from openpyxl import load_workbook

# Load Excel file
wb = load_workbook("option1_job_links.xlsx")
sheet = wb.active

urls = []

# Read URLs (skip header)
for row in sheet.iter_rows(min_row=2, values_only=True):
    url = row[0]
    if url:
        urls.append(url)

print(f"Total URLs: {len(urls)}")

with sync_playwright() as p:
    browser = p.chromium.launch(headless=False)

    for i, url in enumerate(urls):
        try:
            page = browser.new_page()
            print(f"Opening: {url}")

            page.goto(url, timeout=60000)
            page.wait_for_timeout(3000)

            filename = f"screenshot_{i+1}.png"
            page.screenshot(path=filename, full_page=True)

            print(f"Saved: {filename}")

            page.close()

        except Exception as e:
            print(f"Error in {url}: {e}")

    browser.close()